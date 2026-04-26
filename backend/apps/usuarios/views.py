import logging
import os
import re
import secrets
from datetime import timedelta

from django.conf import settings
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.utils import timezone
from decouple import config as env_config

from rest_framework import generics, status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny, IsAdminUser, IsAuthenticated

from apps.usuarios.authentication import UsuarioJWTAuthentication
from apps.usuarios.permissions import EsAdministrador
from apps.usuarios.utils import generar_contrasena, enviar_contrasena_bienvenida
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenBlacklistView
from rest_framework_simplejwt.tokens import RefreshToken

from apps.bitacora.models import BitacoraSistema
from apps.bitacora.utils import registrar_evento
from apps.roles.models import Rol
from apps.usuarios.models import TokenRecuperacion, Usuario, UsuarioRol
from apps.usuarios.serializers import (
    CambiarContrasenaSerializer,
    OlvidarContrasenaSerializer,
    RecuperarContrasenaSerializer,
    UsuarioSerializer,
    UsuarioUpdateSerializer,
)

logger = logging.getLogger(__name__)


class UsuarioProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        usuario = request.user
        serializer = UsuarioSerializer(usuario)
        return Response(serializer.data)

    def patch(self, request):
        usuario = request.user
        serializer = UsuarioUpdateSerializer(
            usuario, data=request.data, partial=True, context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response(UsuarioSerializer(usuario).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



# LoginView: View personalizada para autenticación JWT (BE-02, BE-03)
# Maneja el inicio de sesión retornando tokens JWT
# Permite acceso sin token (AllowAny) ya que es el proceso de login
class LoginView(APIView):
    # permission_classes: Permite acceso público para poder hacer login
    permission_classes = [AllowAny]

    def post(self, request):
        """
        Endpoint POST /api/auth/login/
        Proceso de autenticación:
        1. Recibe correo y contraseña en body JSON
        2. Valida credenciales contra modelo Usuario
        3. Retorna access + refresh tokens JWT
        """
        # Extrae credenciales del request JSON
        # .strip() elimina espacios en blanco al inicio/final del correo
        correo    = request.data.get('correo', '').strip()
        contrasena = request.data.get('contrasena', '')

        # Valida que ambos campos existan
        if not correo or not contrasena:
            return Response(
                {'detail': 'correo y contrasena son requeridos.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Busca usuario en la base de datos por correo
        # Modelo personalizado: apps.usuarios.models.Usuario
        try:
            usuario = Usuario.objects.get(correo=correo)
        except Usuario.DoesNotExist:
            usuario = None

        # check_password() es el método de AbstractBaseUser que compara la contraseña
        # recibida (en texto plano) contra el hash guardado en el campo 'password'.
        # Antes se llamaba a la función suelta check_password(contrasena, usuario.contrasena_hash).
        # Ahora el propio objeto sabe dónde está su hash, por eso se llama sobre la instancia.
        if usuario is None or not usuario.check_password(contrasena):
            # Registra intento fallido en bitácora para auditoría
            registrar_evento(
                request=request,
                accion=BitacoraSistema.Accion.ACCESS_DENIED,
                modulo='autenticacion',
                descripcion=f'Intento de login fallido para correo: {correo}',
            )
            return Response(
                {'detail': 'Credenciales inválidas.'},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        # Verifica estado del usuario (BE-02)
        # Solo usuarios activos pueden iniciar sesión
        if usuario.estado != Usuario.Estado.ACTIVO:
            registrar_evento(
                request=request,
                accion=BitacoraSistema.Accion.ACCESS_DENIED,
                modulo='autenticacion',
                descripcion=f'Login denegado, usuario inactivo: {correo}',
            )
            return Response(
                {'detail': 'Usuario inactivo.'},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        # Genera tokens JWT (BE-03)
        # RefreshToken: Clase de SimpleJWT que crea par access+refresh
        refresh = RefreshToken()
        
        # Personalización del payload del token (BE-03)
        # Agrega datos del usuario al token para evitar llamadas adicionales
        # El frontend puede leer estos datos del access token directamente
        # Campos incluios en el token:
        refresh['user_id']   = usuario.id      # ID único del usuario
        refresh['nombre']   = usuario.nombre # Nombre para mostrar en UI
        refresh['correo']  = usuario.correo  # Correo del usuario
        refresh['tipo_rol'] = usuario.tipo_rol # Rol: admin, docente, estudiante

        # Asigna usuario al request para uso en permisos/logging
        request.usuario = usuario
        # Registra login exitoso en bitácora
        registrar_evento(
            request=request,
            accion=BitacoraSistema.Accion.LOGIN,
            modulo='autenticacion',
            descripcion=f'Login efectivo: {correo}',
        )

        # Retorna tokens al cliente
        # access: Token corto (1 hora) para solicitudes API
        # refresh: Token largo (7 días) para obtener nuevos access
        return Response({
            'access':  str(refresh.access_token),
            'refresh': str(refresh),
        }, status=status.HTTP_200_OK)


class UsuarioUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_usuario_o_error(self, request, pk):
        es_admin = request.user.tipo_rol == Usuario.TipoRol.ADMINISTRADOR
        if not es_admin and request.user.pk != pk:
            return None, Response(
                {'detail': 'No tienes permiso para acceder a este perfil.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        try:
            return Usuario.objects.get(pk=pk), None
        except Usuario.DoesNotExist:
            return None, Response({'detail': 'Usuario no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    def get(self, request, pk):
        usuario, error = self._get_usuario_o_error(request, pk)
        if error:
            return error
        return Response(UsuarioSerializer(usuario).data)

    def patch(self, request, pk):
        usuario, error = self._get_usuario_o_error(request, pk)
        if error:
            return error
        serializer = UsuarioUpdateSerializer(
            usuario, data=request.data, partial=True, context={'request': request}
        )
        if serializer.is_valid():
            campos_modificados = list(serializer.validated_data.keys())
            serializer.save()
            registrar_evento(
                request=request,
                accion=BitacoraSistema.Accion.UPDATE,
                modulo='usuarios',
                descripcion=(
                    f'Perfil actualizado: ID={usuario.id}, correo={usuario.correo}. '
                    f'Campos modificados: {", ".join(campos_modificados)}. '
                    f'Editado por: {request.user.correo}'
                ),
            )
            return Response(UsuarioSerializer(usuario).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UsuarioCreateView(generics.ListCreateAPIView):
    queryset             = Usuario.objects.all().order_by('-fecha_creacion')
    serializer_class     = UsuarioSerializer
    authentication_classes = [UsuarioJWTAuthentication]
    permission_classes   = [EsAdministrador]

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return UsuarioSerializer
        return UsuarioSerializer

    def perform_create(self, serializer):
        contrasena = generar_contrasena()
        serializer.validated_data['contrasena'] = contrasena
        usuario_creado = serializer.save()

        email_enviado = True
        try:
            enviar_contrasena_bienvenida(usuario_creado.nombre, usuario_creado.correo, contrasena)
        except Exception as exc:
            logger.error('Error al enviar correo de bienvenida a %s: %s', usuario_creado.correo, exc)
            email_enviado = False

        registrar_evento(
            request=self.request,
            accion=BitacoraSistema.Accion.CREATE,
            modulo='usuarios',
            descripcion=(
                f'Usuario creado: ID={usuario_creado.id}, correo={usuario_creado.correo}. '
                f'Correo enviado: {email_enviado}'
            ),
        )
        self._email_enviado = email_enviado

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        data = serializer.data
        data_out = dict(data)
        data_out['email_enviado'] = getattr(self, '_email_enviado', True)
        headers = self.get_success_headers(serializer.data)
        return Response(data_out, status=status.HTTP_201_CREATED, headers=headers)


class LogoutView(TokenBlacklistView):
    """Vista personalizada para cerrar sesión y registrar el evento en bitácora."""
    def post(self, request, *args, **kwargs):
        user = request.user
        if user and user.is_authenticated:
            registrar_evento(
                request=request,
                accion=BitacoraSistema.Accion.LOGOUT,
                modulo='autenticacion',
                descripcion=f'Logout efectivo: {user.correo}',
            )
        return super().post(request, *args, **kwargs)


class CambiarContrasenaView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = CambiarContrasenaSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        usuario = request.user
        password_actual  = serializer.validated_data['password_actual']
        nueva_contrasena = serializer.validated_data['nueva_contrasena']

        if not usuario.check_password(password_actual):
            return Response(
                {'password_actual': 'La contraseña actual es incorrecta.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        usuario.set_password(nueva_contrasena)
        usuario.save()

        registrar_evento(
            request=request,
            accion=BitacoraSistema.Accion.UPDATE,
            modulo='autenticacion',
            descripcion=f'Cambio de contraseña: {usuario.correo}',
        )

        return Response(
            {'mensaje': 'Contraseña actualizada correctamente.'},
            status=status.HTTP_200_OK,
        )


_RESPUESTA_GENERICA = {
    'mensaje': 'Si el correo está registrado, recibirás un enlace en los próximos minutos.'
}


class OlvidarContrasenaView(APIView):
    permission_classes = [AllowAny]

    _RATE_LIMIT_MAX     = 1
    _RATE_LIMIT_SECONDS = 60  # 1 minuto entre intentos

    def post(self, request):
        from django.core.cache import cache

        serializer = OlvidarContrasenaSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        correo = serializer.validated_data['correo'].lower()

        # Rate limit por correo: máximo 3 solicitudes cada 15 minutos.
        # Responde igual aunque esté bloqueado para no revelar información.
        cache_key = f'recuperacion_{correo}'
        intentos  = cache.get(cache_key, 0)
        if intentos >= self._RATE_LIMIT_MAX:
            return Response(_RESPUESTA_GENERICA, status=status.HTTP_200_OK)
        cache.set(cache_key, intentos + 1, timeout=self._RATE_LIMIT_SECONDS)

        try:
            usuario = Usuario.objects.get(correo=correo, estado=Usuario.Estado.ACTIVO)
        except Usuario.DoesNotExist:
            # Respuesta genérica: no revelar si el correo existe en el sistema
            return Response(_RESPUESTA_GENERICA, status=status.HTTP_200_OK)

        # Invalidar tokens anteriores pendientes del mismo usuario
        TokenRecuperacion.objects.filter(usuario=usuario, usado=False).update(usado=True)

        token_str = secrets.token_urlsafe(48)
        TokenRecuperacion.objects.create(
            usuario=usuario,
            token=token_str,
            expiracion=timezone.now() + timedelta(minutes=30),
        )

        frontend_url = env_config('FRONTEND_URL', default='http://localhost:5173')
        enlace = f'{frontend_url}/recuperar-contrasena?token={token_str}'

        send_mail(
    subject='Recuperación de contraseña - ABP Platform',
    message=f'Hola {usuario.nombre}, restablece tu contraseña aquí: {enlace}',  # fallback texto plano
    from_email=settings.DEFAULT_FROM_EMAIL,
    recipient_list=[correo],
    fail_silently=False,
    html_message=f"""
    <div style="font-family: Arial, sans-serif; max-width: 480px; margin: 0 auto; padding: 32px; background: #f8f9fa; border-radius: 12px;">
        <h2 style="color: #191c1d; margin-bottom: 8px;">Recupera tu contraseña</h2>
        <p style="color: #5b403d; font-size: 14px;">Hola <strong>{usuario.nombre}</strong>,</p>
        <p style="color: #5b403d; font-size: 14px;">Recibimos una solicitud para restablecer tu contraseña en <strong>ABP Platform</strong>. Haz clic en el botón para continuar:</p>

        <div style="text-align: center; margin: 32px 0;">
            <a href="{enlace}" style="
                background: linear-gradient(135deg, #af101a, #d32f2f);
                color: white;
                padding: 14px 32px;
                border-radius: 8px;
                text-decoration: none;
                font-weight: bold;
                font-size: 15px;
                display: inline-block;
            ">
                Restablecer contraseña
            </a>
        </div>

        <p style="color: #888; font-size: 12px; text-align: center;">Este enlace expira en <strong>30 minutos</strong> y es de un solo uso.</p>
        <p style="color: #888; font-size: 12px; text-align: center;">Si no solicitaste esto, ignora este correo.</p>

        <hr style="border: none; border-top: 1px solid #e0e0e0; margin: 24px 0;">
        <p style="color: #aaa; font-size: 11px; text-align: center;">ABP Platform · UFPS</p>
    </div>
    """,
)

        return Response(_RESPUESTA_GENERICA, status=status.HTTP_200_OK)


class RecuperarContrasenaView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RecuperarContrasenaSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        token_str        = serializer.validated_data['token']
        nueva_contrasena = serializer.validated_data['nueva_contrasena']

        try:
            token_obj = TokenRecuperacion.objects.select_related('usuario').get(token=token_str)
        except TokenRecuperacion.DoesNotExist:
            return Response(
                {'detail': 'Token inválido o expirado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not token_obj.esta_vigente():
            return Response(
                {'detail': 'Token inválido o expirado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        usuario = token_obj.usuario
        usuario.set_password(nueva_contrasena)
        usuario.save()

        token_obj.usado = True
        token_obj.save()

        return Response(
            {'mensaje': 'Contraseña actualizada correctamente.'},
            status=status.HTTP_200_OK,
        )


ROLES_VALIDOS = ['administrador', 'director', 'docente', 'lider_equipo', 'estudiante']


class CargaMasivaEstudiantesView(APIView):
    """
    POST /api/usuarios/carga-masiva/
    Recibe un archivo .xlsx y un campo 'rol', crea usuarios con ese rol.
    Columnas esperadas: nombre, apellido, correo (y opcionalmente codigo_estudiante).
    """
    authentication_classes = [UsuarioJWTAuthentication]
    permission_classes     = [EsAdministrador]
    parser_classes         = [MultiPartParser]

    def post(self, request):
        archivo = request.FILES.get('archivo')
        if archivo is None:
            return Response(
                {'detail': 'Debe adjuntar el archivo en el campo "archivo".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rol = request.data.get('rol', '').strip()
        if rol not in ROLES_VALIDOS:
            return Response(
                {'detail': f'Rol inválido. Opciones: {ROLES_VALIDOS}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        nombre_archivo = (archivo.name or '').lower()
        if not nombre_archivo.endswith('.xlsx'):
            return Response(
                {'detail': 'El archivo debe ser formato .xlsx.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError:
            return Response(
                {'detail': 'Dependencia openpyxl no instalada en el servidor.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
        except InvalidFileException:
            return Response(
                {'detail': 'El archivo .xlsx es inválido o está corrupto.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception:
            return Response(
                {'detail': 'No fue posible leer el archivo .xlsx.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))

        if len(filas) <= 1:
            return Response(
                {'detail': 'El archivo está vacío o no contiene filas de datos.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rol_obj = Rol.objects.filter(nombre__iexact=rol).first()

        creados = 0
        omitidos = 0
        errores = []
        correos_archivo = set()

        for idx, fila in enumerate(filas[1:], start=2):
            fila = list(fila) + [None] * (4 - len(fila)) if len(fila) < 4 else fila
            codigo_estudiante, nombre, apellido, correo = fila[0], fila[1], fila[2], fila[3]

            codigo_estudiante = str(codigo_estudiante).strip() if codigo_estudiante is not None else ''
            nombre   = str(nombre).strip() if nombre is not None else ''
            apellido = str(apellido).strip() if apellido is not None else ''
            correo   = str(correo).strip().lower() if correo is not None else ''

            if not any([codigo_estudiante, nombre, apellido, correo]):
                continue

            if not correo:
                omitidos += 1
                errores.append({'fila': idx, 'correo': correo, 'motivo': 'correo vacío'})
                continue

            try:
                validate_email(correo)
            except ValidationError:
                omitidos += 1
                errores.append({'fila': idx, 'correo': correo, 'motivo': 'correo con formato inválido'})
                continue

            if not nombre or not apellido:
                omitidos += 1
                errores.append({'fila': idx, 'correo': correo, 'motivo': 'nombre o apellido vacío'})
                continue

            if correo in correos_archivo:
                omitidos += 1
                errores.append({'fila': idx, 'correo': correo, 'motivo': 'correo duplicado en el archivo'})
                continue

            if Usuario.objects.filter(correo=correo).exists():
                omitidos += 1
                errores.append({'fila': idx, 'correo': correo, 'motivo': 'correo ya registrado'})
                continue

            correos_archivo.add(correo)
            contrasena = generar_contrasena()

            try:
                usuario = Usuario.objects.create(
                    codigo_estudiante=codigo_estudiante if codigo_estudiante else None,
                    nombre=nombre,
                    apellido=apellido,
                    correo=correo,
                    password=make_password(contrasena),
                    tipo_rol=rol,
                    estado=Usuario.Estado.ACTIVO,
                    is_staff=False,
                    is_superuser=False,
                )
            except Exception as exc:
                omitidos += 1
                errores.append({'fila': idx, 'correo': correo, 'motivo': f'error al crear: {exc}'})
                continue

            if rol_obj is not None:
                UsuarioRol.objects.get_or_create(usuario=usuario, rol=rol_obj)

            try:
                enviar_contrasena_bienvenida(nombre, correo, contrasena)
            except Exception as exc:
                logger.error('Error al enviar correo a %s: %s', correo, exc)
                errores.append({'fila': idx, 'correo': correo, 'motivo': f'usuario creado, pero fallo el correo: {exc}'})

            creados += 1

        registrar_evento(
            request=request,
            accion=BitacoraSistema.Accion.CREATE,
            modulo='usuarios',
            descripcion=(
                f'Carga masiva ({rol}): creados={creados}, '
                f'omitidos={omitidos}, archivo={archivo.name}'
            ),
        )

        return Response(
            {'creados': creados, 'omitidos': omitidos, 'errores': errores},
            status=status.HTTP_200_OK,
        )


class SubirFotoPerfilView(APIView):
    """
    POST /api/usuarios/foto-perfil/
    Sube una imagen desde el dispositivo y la guarda como foto de perfil.
    Campo del form: 'foto' (archivo de imagen).
    """
    authentication_classes = [UsuarioJWTAuthentication]
    permission_classes     = [IsAuthenticated]
    parser_classes         = [MultiPartParser]

    TIPOS_PERMITIDOS = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
    TAMANO_MAX       = 5 * 1024 * 1024  # 5 MB

    def post(self, request):
        archivo = request.FILES.get('foto')
        if not archivo:
            return Response(
                {'detail': 'No se recibió ningún archivo. Usa el campo "foto".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if archivo.content_type not in self.TIPOS_PERMITIDOS:
            return Response(
                {'detail': 'Solo se permiten imágenes JPG, PNG, GIF o WebP.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if archivo.size > self.TAMANO_MAX:
            return Response(
                {'detail': 'La imagen no puede superar 5 MB.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        usuario = request.user
        ext = (archivo.name or 'foto').rsplit('.', 1)[-1].lower()
        if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
            ext = 'jpg'
        nombre_archivo = f'{usuario.id}_{int(timezone.now().timestamp())}.{ext}'

        carpeta = os.path.join(settings.MEDIA_ROOT, 'fotos_perfil')
        os.makedirs(carpeta, exist_ok=True)
        ruta_disco = os.path.join(carpeta, nombre_archivo)

        # Eliminar foto anterior si es un archivo local
        foto_anterior = usuario.foto_perfil or ''
        if 'fotos_perfil' in foto_anterior:
            nombre_anterior = foto_anterior.rsplit('fotos_perfil/', 1)[-1].split('?')[0]
            ruta_anterior = os.path.join(carpeta, nombre_anterior)
            if os.path.isfile(ruta_anterior):
                try:
                    os.remove(ruta_anterior)
                except OSError:
                    pass

        with open(ruta_disco, 'wb') as f:
            for chunk in archivo.chunks():
                f.write(chunk)

        url_relativa = f'{settings.MEDIA_URL}fotos_perfil/{nombre_archivo}'

        usuario.foto_perfil = url_relativa
        usuario.save(update_fields=['foto_perfil', 'fecha_actualizacion'])

        registrar_evento(
            request=request,
            accion=BitacoraSistema.Accion.UPDATE,
            modulo='usuarios',
            descripcion=f'Foto de perfil actualizada: ID={usuario.id}, correo={usuario.correo}',
        )

        return Response({'foto_perfil': url_relativa}, status=status.HTTP_200_OK)